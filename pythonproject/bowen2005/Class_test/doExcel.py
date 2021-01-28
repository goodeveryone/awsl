
import openpyxl

# 创建一个工作薄对象
wb = openpyxl.Workbook()
# 创建一个标签页
wb.create_sheet('test_case',1)
# 保存为一个xlsx格式的文件
wb.save("ceshi.xlsx")

# 读取excel中的数据
# 打开工作薄
# wb = openpyxl.load_workbook('ceshi.xlsx')
# 选取标签页
# sh = wb['Sheet']
# 单元格读取数据
#     # 参数:
#         # 1. row 行
#         # 2. col 列
# ce = sh.cell(1,1)
# print(ce.value)
# 按行读取数据
# for case in list(sh.rows):
#     case_id = case[0].value
#     case_excepted = case[1].value
#     case_data = case[2].value
#     print(case_id,case_excepted,case_data)
# 按列读取数据
# columns_data = sh.columns
# print(list(columns_data))
# 写入数据
# 第五行第一列
# sh.cell(row=5,column=1,value='result')
# # 保存
# wb.save('ceshi.xlsx')
# 获取最大行
# print(sh.max_row)
# # 获取最大列
# print(sh.max_column)
# 创建表单 名字和索引指定位置0,1,2
# wb.create_sheet('test_case',1)
# # 获取所有的标签页名
# # print(wb.sheetnames)
#
# wb.save('ceshi.xlsx')
# # 关闭工作薄
# wb.close()
# a = isinstance('',str)
# print(a)

# def hanshu(user,password):
#     print(user['data'],password['data'])
#
# a = [{'data':(1,2)},{'data':(1,2)}]
# hanshu(*a)