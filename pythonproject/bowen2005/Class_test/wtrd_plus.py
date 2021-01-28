import openpyxl
# f = openpyxl.Workbook()
# f.create_sheet("sheet_test")
# f.save("ceshi_plus.xlsx")

f = openpyxl.load_workbook('ceshi_plus.xlsx')
a = f['sheet']
b = a.cell(1,1)
print(b)



















































